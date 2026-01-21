import io
from typing import Optional
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from app.models.rfq import RFQ, RFQVersion

def register_fonts():
    """Attempt to register a CJK font."""
    # List of candidate font paths (Mac, generic Linux, etc.)
    font_paths = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
    ]
    
    available_font = "Helvetica" # Default
    
    for path in font_paths:
        try:
            # Note: reportlab might need 'tc' index for ttc files
            # Just try registering it as a name
            font_name = "CJKFont"
            pdfmetrics.registerFont(TTFont(font_name, path))
            available_font = font_name
            break
        except Exception:
            continue
            
    return available_font

def generate_pdf(rfq: RFQ, version: RFQVersion) -> io.BytesIO:
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Fonts
    font_reg = register_fonts()
    p.setFont(font_reg, 12)
    
    # 1. Header
    p.setFont(font_reg, 24)
    p.drawString(2*cm, height - 3*cm, f"報價單")
    
    p.setFont(font_reg, 10)
    p.drawString(2*cm, height - 4*cm, f"單號: {rfq.rfq_no}")
    if version.project_name:
        p.drawString(2*cm, height - 4.5*cm, f"專案: {version.project_name}")

    p.drawString(12*cm, height - 4*cm, f"日期: {version.created_at.strftime('%Y-%m-%d')}")
    p.drawString(12*cm, height - 4.5*cm, f"版本: v{version.version_number}")

    # 2. Vendor Info (Snapshot)
    v_snap = version.vendor_snapshot
    if v_snap:
        p.drawString(2*cm, height - 6*cm, f"廠商: {v_snap.get('company_name', '')}")
        p.drawString(2*cm, height - 6.5*cm, f"聯絡人: {v_snap.get('primary_contact_name', '')}")
        p.drawString(2*cm, height - 7*cm, f"信箱: {v_snap.get('email', '')}")

    # 3. Items Table Header
    y_pos = height - 9*cm
    p.line(2*cm, y_pos - 0.2*cm, 19*cm, y_pos - 0.2*cm)
    
    p.drawString(2*cm, y_pos, "項目名稱")
    p.drawString(8*cm, y_pos, "規格")
    p.drawString(12*cm, y_pos, "數量")
    p.drawString(14*cm, y_pos, "單價")
    p.drawString(17*cm, y_pos, "金額")
    
    y_pos -= 1*cm
    
    # 4. Items
    total = 0
    for item in version.items:
        # Simple text wrapping could be added here, but for now just truncate or overlap
        name = item.name or ""
        p.drawString(2*cm, y_pos, name[:20]) # Limit chars
        
        spec = item.description or ""
        p.drawString(8*cm, y_pos, spec[:15])
        
        qty_str = f"{item.quantity} {item.unit}"
        p.drawString(12*cm, y_pos, qty_str)
        
        price = item.unit_price or 0
        p.drawString(14*cm, y_pos, f"{price:,.0f}")
        
        subtotal = item.line_subtotal or 0
        p.drawString(17*cm, y_pos, f"{subtotal:,.0f}")
        
        total += subtotal
        y_pos -= 0.8*cm
        
        if y_pos < 3*cm:
            p.showPage()
            p.setFont(font_reg, 10)
            y_pos = height - 3*cm
            
    # 5. Footer
    p.line(2*cm, y_pos + 0.5*cm, 19*cm, y_pos + 0.5*cm)
    p.setFont(font_reg, 12)
    p.drawString(14*cm, y_pos - 1*cm, f"總計: {total:,.0f}")
    
    p.setFont(font_reg, 10)
    if version.notes:
        p.drawString(2*cm, y_pos - 3*cm, f"備註: {version.notes}")

    p.save()
    buffer.seek(0)
    return buffer

def generate_excel(rfq: RFQ, version: RFQVersion) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報價單"
    
    # Headers
    ws['A1'] = "報價單"
    ws['A1'].font = Font(size=20, bold=True)
    
    ws['A3'] = "單號:"
    ws['B3'] = rfq.rfq_no
    
    ws['A4'] = "專案:"
    ws['B4'] = version.project_name
    
    ws['D3'] = "日期:"
    ws['E3'] = version.created_at.strftime('%Y-%m-%d')
    
    # Table Header
    headers = ["項目名稱", "規格", "數量", "單位", "單價", "金額"]
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=7, column=col_num)
        c.value = header
        c.font = Font(bold=True)
        c.border = Border(bottom=Side(style='thin'))
        
    # Items
    row_num = 8
    for item in version.items:
        ws.cell(row=row_num, column=1, value=item.name)
        ws.cell(row=row_num, column=2, value=item.description)
        ws.cell(row=row_num, column=3, value=item.quantity)
        ws.cell(row=row_num, column=4, value=item.unit)
        ws.cell(row=row_num, column=5, value=item.unit_price)
        ws.cell(row=row_num, column=6, value=item.line_subtotal)
        row_num += 1
        
    # Total
    ws.cell(row=row_num+1, column=5, value="總計")
    ws.cell(row=row_num+1, column=6, value=version.total_amount)
    
    # Adjust widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
