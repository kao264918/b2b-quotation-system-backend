
"""
Quote Export Service - PDF and Excel generation for quotes
"""
import io
import os
import subprocess
import tempfile
from copy import copy
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from app.models.quote import Quote

ITEM_START_ROW = 17
ITEM_END_ROW = 48
ITEM_BLOCK_HEIGHT = 2
BASE_ITEM_CAPACITY = (ITEM_END_ROW - ITEM_START_ROW + 1) // ITEM_BLOCK_HEIGHT
BASE_TOTAL_ROW = 49
ITEM_CLEAR_MAX_COL = 14
PROTOTYPE_ITEM_TOP_ROW = 17


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int = ITEM_CLEAR_MAX_COL) -> None:
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        target_cell._style = copy(source_cell._style)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format
        target_cell.value = None

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _extract_and_unmerge_ranges(ws, start_row: int):
    ranges_to_shift = [rng for rng in list(ws.merged_cells.ranges) if rng.min_row >= start_row]
    shifted_specs = [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in ranges_to_shift
    ]
    for rng in ranges_to_shift:
        ws.unmerge_cells(str(rng))
    return shifted_specs


def _reapply_shifted_merged_ranges(ws, merge_specs, row_offset: int) -> None:
    for min_row, min_col, max_row, max_col in merge_specs:
        ws.merge_cells(
            start_row=min_row + row_offset,
            start_column=min_col,
            end_row=max_row + row_offset,
            end_column=max_col,
        )


def _ensure_item_block_merges(ws, top_row: int) -> None:
    merge_specs = [
        (top_row, 5, top_row + 1, 5),
        (top_row, 6, top_row + 1, 6),
        (top_row, 7, top_row, 9),
        (top_row + 1, 7, top_row + 1, 9),
        (top_row, 10, top_row + 1, 10),
        (top_row, 11, top_row + 1, 11),
        (top_row, 12, top_row + 1, 12),
    ]

    existing_ranges = {str(rng) for rng in ws.merged_cells.ranges}
    for start_row, start_col, end_row, end_col in merge_specs:
        coord = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
        if coord not in existing_ranges:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _normalize_item_block_display_styles(ws, top_row: int) -> None:
    display_style_map = {
        (top_row, 5): (PROTOTYPE_ITEM_TOP_ROW, 5),
        (top_row, 6): (PROTOTYPE_ITEM_TOP_ROW, 6),
        (top_row, 7): (PROTOTYPE_ITEM_TOP_ROW, 7),
        (top_row, 10): (PROTOTYPE_ITEM_TOP_ROW, 10),
        (top_row, 11): (PROTOTYPE_ITEM_TOP_ROW, 11),
        (top_row, 12): (PROTOTYPE_ITEM_TOP_ROW, 12),
        (top_row + 1, 7): (PROTOTYPE_ITEM_TOP_ROW + 1, 7),
    }

    for (target_row, target_col), (source_row, source_col) in display_style_map.items():
        source_cell = ws.cell(row=source_row, column=source_col)
        target_cell = ws.cell(row=target_row, column=target_col)
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def _find_last_used_row(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return row
    return 1


def generate_quote_pdf(quote: Quote, version: Optional[int] = None) -> bytes:
    """
    Generate a PDF for the given quote.
    Uses LibreOffice (soffice) to convert the generated Excel file to PDF.
    """
    # 1. Generate Excel first (Byte Stream)
    excel_stream = generate_quote_excel_stream(quote, version)
    
    # 2. Save to temporary file
    with tempfile.TemporaryDirectory() as tmpdirname:
        xlsx_path = os.path.join(tmpdirname, "temp_quote.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(excel_stream.getvalue())
            
        # 3. Run LibreOffice conversion
        # Command: soffice --headless --convert-to pdf --outdir <dir> <file>
        # Adjust path for macOS standard install
        soffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if not os.path.exists(soffice_path):
             # Fallback check for linux/standard path or alias
             soffice_path = "soffice" 

        cmd = [
            soffice_path,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            tmpdirname,
            xlsx_path
        ]
        
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            # 4. Read the resulting PDF
            # LibreOffice keeps the same basename: temp_quote.pdf
            pdf_path = os.path.join(tmpdirname, "temp_quote.pdf")
            if os.path.exists(pdf_path):
                with open(pdf_path, "rb") as f:
                    pdf_content = f.read()
                return pdf_content
            else:
                # Fallback or error?
                # If PDF generation fails, we might throw an error/log it
                raise Exception("PDF file was not generated by LibreOffice")
                
        except subprocess.CalledProcessError as e:
            print(f"LibreOffice conversion failed: {e.stderr.decode()}")
            raise Exception("Failed to convert Excel to PDF. Ensure LibreOffice is installed.")


def generate_quote_excel(quote: Quote, version: Optional[int] = None) -> bytes:
    """Wrapper to return bytes for the router."""
    stream = generate_quote_excel_stream(quote, version)
    return stream.getvalue()


def generate_quote_excel_stream(quote: Quote, version_num: Optional[int] = None) -> io.BytesIO:
    """
    Generate an Excel file for the given quote using the template.
    """
    # Load template
    # Assuming app/assets/quote_template.xlsx exists relative to project root
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(base_dir, "app", "assets", "quote_template.xlsx")
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading template: {e}")
        # Fallback to empty if template missing (should not happen in prod if deployed correctly)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "報價單"

    # Remove extra sheets if any
    if len(wb.sheetnames) > 1:
        for sheet_name in wb.sheetnames:
            if sheet_name != ws.title:
                del wb[sheet_name]

    # Rename active sheet: "MMDD CustomerName"
    # Format MMDD
    ver = version_num or quote.version
    dt = quote.created_at
    mmdd = dt.strftime('%m%d')
    cust_name = quote.customer.company_name if quote.customer else "Unknown"
    
    # Sanitize sheet name (remove invalid chars: \ * ? : [ ] /)
    import re
    safe_cust_name = re.sub(r'[\\*?:/\[\]]', '_', cust_name)
    
    # Excel sheet name limit is 31 chars. e.g. "0129 ClientName v1"
    new_sheet_name = f"{mmdd} {safe_cust_name} v{ver}"
    ws.title = new_sheet_name[:31]

    # Helper to safely set cell value (handles merged cells)
    def safe_write(arg1, arg2, arg3=None):
        if isinstance(arg1, int):
            row, col, value = arg1, arg2, arg3
            cell = ws.cell(row=row, column=col)
        else:
            coord, value = arg1, arg2
            cell = ws[coord]
        
        # Check if cell is merged
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
             for merged_range in ws.merged_cells.ranges:
                 if merged_range.min_row == cell.row and merged_range.min_col == cell.column:
                     # Top-left of merge
                     cell.value = value
                     return
                 elif cell.coordinate in merged_range:
                     # Find top left
                     top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                     top_left_cell.value = value
                     return
        else:
             cell.value = value
    
    # 1. Header Info
    # J5: Date
    safe_write('J5', quote.created_at.strftime('%Y/%m/%d'))
    
    # J6-J9: Customer Info
    if quote.customer:
        safe_write('J6', quote.customer.company_name)
        safe_write('J7', quote.customer.tax_id or "")
        safe_write('J8', quote.customer.contact_phone or "")
        safe_write('J9', quote.customer.contact_name or "")
        
    # L1: Quote No
    safe_write('L1', quote.quote_number)
    
    # I2: Title / Project Name
    safe_write('I2', quote.title)
    # J10: Template header field "項目名稱" should mirror quote title
    safe_write('J10', quote.title)

    # 2. Overwrite Table Headers (Row 16)
    # The template might have generic headers. We enforce ours.
    # E(5): Item No
    # F(6): Unit (replacing Type since QuoteItem has no type but has Unit)
    # G(7): Description (Merged G-I)
    # J(10): Qty
    # K(11): Unit Price
    # L(12): Amount
    
    safe_write(16, 5, "項次")
    safe_write(16, 6, "單位") 
    safe_write(16, 7, "項目說明")
    safe_write(16, 10, "數量")
    safe_write(16, 11, "單價")
    safe_write(16, 12, "金額")
    
    # 3. Items Logic
    required_blocks = len(quote.items)
    visible_block_count = max(BASE_ITEM_CAPACITY, required_blocks)
    extra_blocks = max(0, required_blocks - BASE_ITEM_CAPACITY)
    extra_rows = extra_blocks * ITEM_BLOCK_HEIGHT

    if extra_rows > 0:
        merge_specs = _extract_and_unmerge_ranges(ws, BASE_TOTAL_ROW)
        ws.insert_rows(BASE_TOTAL_ROW, extra_rows)
        _reapply_shifted_merged_ranges(ws, merge_specs, extra_rows)

        for block_index in range(extra_blocks):
            target_top_row = ITEM_START_ROW + (BASE_ITEM_CAPACITY + block_index) * ITEM_BLOCK_HEIGHT
            _copy_row_style(ws, PROTOTYPE_ITEM_TOP_ROW, target_top_row)
            _copy_row_style(ws, PROTOTYPE_ITEM_TOP_ROW + 1, target_top_row + 1)
            _ensure_item_block_merges(ws, target_top_row)

    item_region_end_row = ITEM_START_ROW + visible_block_count * ITEM_BLOCK_HEIGHT - 1
    for r in range(ITEM_START_ROW, item_region_end_row + 1):
        for c in range(1, ITEM_CLEAR_MAX_COL + 1):
            safe_write(r, c, None)

    for block_index, item in enumerate(quote.items):
        top_row = ITEM_START_ROW + block_index * ITEM_BLOCK_HEIGHT
        second_row = top_row + 1
        _ensure_item_block_merges(ws, top_row)
        _normalize_item_block_display_styles(ws, top_row)

        safe_write(top_row, 5, str(block_index + 1))
        safe_write(top_row, 6, item.unit)
        safe_write(top_row, 7, item.name)
        safe_write(second_row, 7, item.description or None)
        safe_write(top_row, 10, float(item.quantity))
        safe_write(top_row, 11, float(item.unit_price))
        safe_write(top_row, 12, float(item.subtotal))

    # 4. Totals (Rows 49, 50, 51)
    # K(11): Label, L(12): Value
    subtotal_row = BASE_TOTAL_ROW + extra_rows
    tax_row = subtotal_row + 1
    grand_total_row = subtotal_row + 2

    def fmt_price(val):
        return f"NT$ {val:,.0f}"

    safe_write(subtotal_row, 11, "小計:")
    safe_write(subtotal_row, 12, fmt_price(quote.subtotal))
    
    safe_write(tax_row, 11, "稅金:")
    safe_write(tax_row, 12, fmt_price(quote.tax_total))
    
    safe_write(grand_total_row, 11, "總計:")
    safe_write(grand_total_row, 12, fmt_price(quote.total))

    last_used_row = _find_last_used_row(ws)
    ws.print_area = f"$B$1:$M${last_used_row}"
    ws.print_title_rows = "$16:$16"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
