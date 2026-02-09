"""
Invoice Export Service - PDF and Excel generation for invoices
"""
from io import BytesIO
from typing import Optional
from decimal import Decimal

from app.models.invoice import Invoice


def generate_invoice_pdf(invoice: Invoice) -> bytes:
    """
    Generate a PDF for the given invoice.
    Uses ReportLab for PDF generation with Traditional Chinese support.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    
    # Try to register a Chinese font (fallback to default if not available)
    try:
        pdfmetrics.registerFont(TTFont('NotoSansTC', '/System/Library/Fonts/PingFang.ttc'))
        chinese_font = 'NotoSansTC'
    except:
        chinese_font = 'Helvetica'
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title', 
        parent=styles['Title'],
        fontName=chinese_font,
        fontSize=18,
        spaceAfter=12
    )
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10
    )
    
    elements = []
    
    # Header
    elements.append(Paragraph(f"請款單 {invoice.invoice_number}", title_style))
    elements.append(Spacer(1, 12))
    
    # Invoice Info
    elements.append(Paragraph(f"<b>狀態：</b>{get_status_label(invoice.status)}", normal_style))
    if invoice.accounting_status:
        elements.append(Paragraph(f"<b>付款狀態：</b>{get_accounting_status_label(invoice.accounting_status)}", normal_style))
    if invoice.issued_at:
        elements.append(Paragraph(f"<b>發出日期：</b>{invoice.issued_at.strftime('%Y-%m-%d')}", normal_style))
    if invoice.due_date:
        elements.append(Paragraph(f"<b>到期日：</b>{invoice.due_date.strftime('%Y-%m-%d')}", normal_style))
    elements.append(Spacer(1, 12))
    
    # Customer Info
    if invoice.customer:
        elements.append(Paragraph("<b>客戶資訊</b>", normal_style))
        elements.append(Paragraph(f"公司：{invoice.customer.company_name}", normal_style))
        if hasattr(invoice.customer, 'tax_id') and invoice.customer.tax_id:
            elements.append(Paragraph(f"統編：{invoice.customer.tax_id}", normal_style))
        elements.append(Spacer(1, 12))
    
    # Items Table
    if invoice.items:
        elements.append(Paragraph("<b>請款明細</b>", normal_style))
        elements.append(Spacer(1, 6))
        
        table_data = [['項目名稱', '單位', '數量', '單價', '小計']]
        for item in invoice.items:
            table_data.append([
                item.name,
                item.unit,
                str(item.quantity),
                f"${item.unit_price:,.2f}",
                f"${item.subtotal:,.2f}"
            ])
        
        table = Table(table_data, colWidths=[150, 40, 50, 70, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), chinese_font),
            ('FONTNAME', (0, 1), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))
    
    # Totals
    elements.append(Paragraph(f"<b>未稅小計：</b> ${invoice.subtotal:,.2f}", normal_style))
    elements.append(Paragraph(f"<b>稅額：</b> ${invoice.tax_total:,.2f}", normal_style))
    elements.append(Paragraph(f"<b>總計：</b> ${invoice.total:,.2f}", normal_style))
    
    # Notes
    if invoice.notes:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("<b>備註：</b>", normal_style))
        elements.append(Paragraph(invoice.notes, normal_style))
    
    doc.build(elements)
    return buffer.getvalue()


def generate_invoice_excel(invoice: Invoice) -> bytes:
    """
    Generate an Excel file for the given invoice.
    Uses openpyxl for Excel generation.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "請款單"
    
    # Header
    ws['A1'] = f"請款單 {invoice.invoice_number}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    # Invoice Info
    ws['A3'] = "狀態"
    ws['B3'] = get_status_label(invoice.status)
    ws['A4'] = "付款狀態"
    ws['B4'] = get_accounting_status_label(invoice.accounting_status) if invoice.accounting_status else "-"
    ws['A5'] = "發出日期"
    ws['B5'] = invoice.issued_at.strftime('%Y-%m-%d') if invoice.issued_at else "-"
    ws['A6'] = "到期日"
    ws['B6'] = invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else "-"
    
    # Customer Info
    if invoice.customer:
        ws['A8'] = "客戶"
        ws['A8'].font = Font(bold=True)
        ws['A9'] = "公司"
        ws['B9'] = invoice.customer.company_name
        start_row = 11
    else:
        start_row = 8
    
    # Items Header
    ws[f'A{start_row}'] = "請款明細"
    ws[f'A{start_row}'].font = Font(bold=True)
    start_row += 1
    
    headers = ['項目名稱', '單位', '數量', '單價', '小計']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))
    
    # Items Data
    for idx, item in enumerate(invoice.items):
        row = start_row + idx + 1
        ws.cell(row=row, column=1, value=item.name)
        ws.cell(row=row, column=2, value=item.unit)
        ws.cell(row=row, column=3, value=float(item.quantity))
        ws.cell(row=row, column=4, value=float(item.unit_price))
        ws.cell(row=row, column=5, value=float(item.subtotal))
        ws.cell(row=row, column=5).number_format = '#,##0.00'
    
    # Totals
    totals_row = start_row + len(invoice.items) + 2
    ws[f'D{totals_row}'] = "未稅小計"
    ws[f'E{totals_row}'] = float(invoice.subtotal)
    ws[f'E{totals_row}'].number_format = '#,##0.00'
    
    ws[f'D{totals_row + 1}'] = "稅額"
    ws[f'E{totals_row + 1}'] = float(invoice.tax_total)
    ws[f'E{totals_row + 1}'].number_format = '#,##0.00'
    
    ws[f'D{totals_row + 2}'] = "總計"
    ws[f'E{totals_row + 2}'] = float(invoice.total)
    ws[f'E{totals_row + 2}'].font = Font(bold=True)
    ws[f'E{totals_row + 2}'].number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def get_status_label(status: str) -> str:
    """Get Chinese label for invoice status."""
    labels = {
        "draft": "草稿",
        "issued": "已發出",
        "paid": "已付款",
        "void": "作廢"
    }
    return labels.get(status, status)


def get_accounting_status_label(status: str) -> str:
    """Get Chinese label for accounting status."""
    labels = {
        "unpaid": "未付款",
        "paid": "已付款"
    }
    return labels.get(status, status)
