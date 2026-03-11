from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.quote_export import generate_quote_excel


def _build_quote(items=None):
    now = datetime(2026, 3, 10, tzinfo=timezone.utc)
    return SimpleNamespace(
        quote_number="QUO-2603-001",
        version=3,
        title="2026 Q2 展場整合報價",
        created_at=now,
        subtotal=Decimal("1000.00"),
        tax_total=Decimal("50.00"),
        total=Decimal("1050.00"),
        items=items or [],
        customer=SimpleNamespace(
            company_name="測試客戶股份有限公司",
            tax_id="12345678",
            contact_phone="0912345678",
            contact_name="王小明",
        ),
    )


def _build_items(count: int):
    return [
        SimpleNamespace(
            name=f"項目{index}",
            description=f"說明{index}",
            quantity=1,
            unit="式",
            unit_price=1000,
            subtotal=1000,
        )
        for index in range(1, count + 1)
    ]


def _font_color_repr(cell):
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    return color.value


def _border_color_repr(cell):
    color = cell.border.bottom.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    return color.value


def test_quote_excel_header_uses_customer_primary_contact_and_title():
    quote = _build_quote()

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert worksheet["H2"].value == "2026 Q2 展場整合報價"
    assert worksheet["J10"].value == "2026 Q2 展場整合報價"
    assert worksheet["J6"].value == "測試客戶股份有限公司"
    assert worksheet["J7"].value == "12345678"
    assert worksheet["J8"].value == "0912345678"
    assert worksheet["J9"].value == "王小明"


def test_quote_excel_preserves_template_footer_notes_and_bank_info():
    quote = _build_quote()

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert str(worksheet["D56"].value).strip() == "備  註"
    assert str(worksheet["F56"].value).strip() == "※此訂單簽訂後與正式合約具同等效力。"
    assert worksheet["B60"].value == "客 戶 回 簽"
    assert worksheet["F65"].value == "匯款銀行"


def test_quote_excel_still_clears_item_table_template_rows():
    quote = _build_quote()

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert worksheet["G17"].value is None
    assert worksheet["F17"].value is None
    assert worksheet["J17"].value is None
    assert worksheet["G18"].value is None


def test_quote_excel_expands_rows_and_preserves_footer_for_large_item_count():
    quote = _build_quote(items=_build_items(18))

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert worksheet["G49"].value == "項目17"
    assert worksheet["G51"].value == "項目18"
    assert worksheet["K53"].value == "小計:"
    assert worksheet["L53"].value == "NT$ 1,000"
    assert str(worksheet["D60"].value).strip() == "備  註"
    assert worksheet["B64"].value == "客 戶 回 簽"
    assert worksheet["F69"].value == "匯款銀行"


def test_quote_excel_normalizes_item_font_color_for_late_rows():
    quote = _build_quote(items=_build_items(18))

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert _font_color_repr(worksheet["G45"]) != "FFFF0000"
    assert _font_color_repr(worksheet["G47"]) != "FFFF0000"
    assert _font_color_repr(worksheet["G49"]) != "FFFF0000"
    assert _font_color_repr(worksheet["G51"]) != "FFFF0000"


def test_quote_excel_updates_print_settings_for_multi_page_output():
    quote = _build_quote(items=_build_items(18))

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    print_area = str(worksheet.print_area)
    assert print_area.startswith("'0310 測試客戶股份有限公司 v3'!$B$1:$M$")
    assert int(print_area.rsplit("$", 1)[-1]) > 76
    assert worksheet.print_title_rows in (None, "")
    assert worksheet.page_setup.fitToWidth == 1
    assert worksheet.page_setup.fitToHeight == 0


def test_quote_excel_emphasizes_grand_total_only():
    quote = _build_quote(items=_build_items(2))

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    assert worksheet["K49"].font.bold is False
    assert worksheet["L49"].font.bold is False
    assert worksheet["K50"].font.bold is False
    assert worksheet["L50"].font.bold is False
    assert worksheet["K51"].font.bold is True
    assert worksheet["L51"].font.bold is True


def test_quote_excel_normalizes_item_bottom_border_for_template_and_expanded_rows():
    quote = _build_quote(items=_build_items(18))

    excel_bytes = generate_quote_excel(quote)
    workbook = load_workbook(BytesIO(excel_bytes))
    worksheet = workbook.active

    rows_to_check = [18, 20, 48, 50, 52]
    cols_to_check = ["E", "F", "G", "H", "I", "J", "K", "L"]

    expected_style = worksheet["F18"].border.bottom.style
    expected_color = _border_color_repr(worksheet["F18"])
    assert expected_style == "thin"
    assert expected_color == "FF000000"

    for row in rows_to_check:
        styles = [worksheet[f"{col}{row}"].border.bottom.style for col in cols_to_check]
        colors = [_border_color_repr(worksheet[f"{col}{row}"]) for col in cols_to_check]
        assert all(style == expected_style for style in styles)
        assert all(color == expected_color for color in colors)
